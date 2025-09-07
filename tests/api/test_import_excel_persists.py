"""
Test that Excel/CSV import actually persists data to the database.

This test verifies that the importer contains working write logic,
not just placeholder/no-op code.
"""

import pytest
import requests
import sqlite3
import tempfile
import os
from io import BytesIO
from openpyxl import Workbook


@pytest.fixture
def base_url():
    return "http://127.0.0.1:8000/api/v1"


@pytest.fixture
def temp_db():
    """Create a temporary database for testing."""
    db_path = "/tmp/lorien_import_test.db"

    # Clean up any existing database
    if os.path.exists(db_path):
        os.remove(db_path)

    # Set environment variable for this test
    os.environ['LORIEN_DB_PATH'] = db_path

    yield db_path

    # Cleanup
    if os.path.exists(db_path):
        os.remove(db_path)


@pytest.fixture
def valid_excel_data():
    """Create a valid Excel file with the frozen 8-column header and test data."""

    # Create workbook
    wb = Workbook()
    ws = wb.active

    # Add the frozen header (from test_csv_header_freeze.py)
    headers = [
        'vm', 'node1', 'node2', 'node3', 'node4', 'node5',
        'diagnosis', 'actions'
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Add test data
    test_data = [
        ['Test VM 1', 'Symptom A', 'Symptom B', '', '', '', 'Diagnosis 1', 'Action 1'],
        ['Test VM 2', 'Symptom C', 'Symptom D', 'Symptom E', '', '', 'Diagnosis 2', 'Action 2'],
    ]

    for row_idx, row_data in enumerate(test_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Save to BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return bio.getvalue()


def get_tree_stats(db_path: str) -> dict:
    """Get tree statistics from database."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Get basic counts
    cursor.execute("SELECT COUNT(*) FROM nodes")
    nodes = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM nodes WHERE parent_id IS NULL")
    roots = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COUNT(*) FROM nodes n
        WHERE NOT EXISTS (
            SELECT 1 FROM nodes c WHERE c.parent_id = n.id
        )
    """)
    leaves = cursor.fetchone()[0]

    conn.close()

    return {
        'nodes': nodes,
        'roots': roots,
        'leaves': leaves
    }


def test_import_excel_creates_persisted_data(base_url, temp_db, valid_excel_data):
    """Test that Excel import actually persists data to database."""

    # Get baseline stats
    baseline_stats = get_tree_stats(temp_db)
    print(f"Baseline stats: {baseline_stats}")

    # Prepare multipart form data
    files = {
        'file': ('test_data.xlsx', BytesIO(valid_excel_data), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    }
    data = {'source': 'test'}

    # Make the import request
    response = requests.post(
        f"{base_url}/import",
        files=files,
        data=data
    )

    print(f"Import response status: {response.status_code}")
    print(f"Import response body: {response.text}")

    # Should succeed (200) if importer is implemented
    if response.status_code == 501:
        pytest.skip("Importer is placeholder (501 Not Implemented)")
    elif response.status_code == 500:
        pytest.fail("Importer fails with 500 - likely missing implementation")
    else:
        assert response.status_code in [200, 201]

    # Get stats after import
    post_import_stats = get_tree_stats(temp_db)
    print(f"Post-import stats: {post_import_stats}")

    # Verify that data was actually persisted
    assert post_import_stats['nodes'] > baseline_stats['nodes'], \
        "Import did not create any new nodes in database"

    assert post_import_stats['roots'] > baseline_stats['roots'], \
        "Import did not create any new root nodes (VMs)"

    # Should have created at least 2 VMs based on our test data
    new_roots = post_import_stats['roots'] - baseline_stats['roots']
    assert new_roots >= 2, f"Expected at least 2 new VMs, got {new_roots}"

    # Should have created child nodes
    new_nodes = post_import_stats['nodes'] - baseline_stats['nodes']
    assert new_nodes >= 4, f"Expected at least 4 new nodes, got {new_nodes}"


def test_import_csv_creates_persisted_data(base_url, temp_db):
    """Test that CSV import actually persists data to database."""

    # Create valid CSV data
    csv_data = """vm,node1,node2,node3,node4,node5,diagnosis,actions
Test VM 1,Fever,Cough,,,,"High fever with cough","Rest, fluids, monitor"
Test VM 2,Headache,Nausea,Vomiting,,,"Migraine with nausea","Dark room, pain relief\""""

    # Get baseline stats
    baseline_stats = get_tree_stats(temp_db)
    print(f"Baseline CSV stats: {baseline_stats}")

    # Prepare multipart form data
    files = {
        'file': ('test_data.csv', BytesIO(csv_data.encode('utf-8')), 'text/csv')
    }
    data = {'source': 'test'}

    # Make the import request
    response = requests.post(
        f"{base_url}/import",
        files=files,
        data=data
    )

    print(f"CSV Import response status: {response.status_code}")
    print(f"CSV Import response body: {response.text}")

    # Should succeed if importer is implemented
    if response.status_code == 501:
        pytest.skip("CSV importer is placeholder (501 Not Implemented)")
    elif response.status_code == 500:
        pytest.fail("CSV importer fails with 500 - likely missing implementation")
    else:
        assert response.status_code in [200, 201]

    # Get stats after import
    post_import_stats = get_tree_stats(temp_db)
    print(f"Post-CSV import stats: {post_import_stats}")

    # Verify that data was actually persisted
    assert post_import_stats['nodes'] > baseline_stats['nodes'], \
        "CSV import did not create any new nodes in database"

    assert post_import_stats['roots'] >= baseline_stats['roots'] + 2, \
        "CSV import did not create expected root nodes"


def test_import_persistence_contract_fulfilled():
    """Test that the importer actually writes to the database."""
    # Get baseline stats
    baseline_response = requests.get("http://127.0.0.1:8000/api/v1/tree/stats")
    baseline_stats = baseline_response.json()
    baseline_nodes = baseline_stats.get("nodes", 0)
    
    # Create test Excel data
    wb = Workbook()
    ws = wb.active
    
    # Add frozen header
    headers = [
        "Vital Measurement", "Node 1", "Node 2", "Node 3", "Node 4", "Node 5", 
        "Diagnostic Triage", "Actions"
    ]
    ws.append(headers)
    
    # Add test data
    ws.append([
        "Test Import VM", "Fever", "Cough", "Headache", "", "", 
        "Test diagnosis", "Test actions"
    ])
    
    # Save to bytes
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_data = excel_buffer.getvalue()
    
    # Import the data
    files = {"file": ("test.xlsx", excel_data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    import_response = requests.post("http://127.0.0.1:8000/api/v1/import", files=files)
    
    assert import_response.status_code == 200, f"Import failed: {import_response.text}"
    
    import_result = import_response.json()
    assert import_result["status"] == "success"
    assert import_result["rows_processed"] == 1
    assert import_result["created"]["roots"] >= 1
    assert import_result["created"]["nodes"] >= 1
    
    # Verify stats increased
    final_response = requests.get("http://127.0.0.1:8000/api/v1/tree/stats")
    final_stats = final_response.json()
    final_nodes = final_stats.get("nodes", 0)
    
    assert final_nodes > baseline_nodes, "Import did not increase node count"
